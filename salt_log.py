from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Color
from openpyxl.styles.colors import YELLOW

from employee import Employee
from datetime import datetime
from week import SaltWeek

from datetime import date
from datetime import datetime
import re

class SaltLog:
    def __init__(self, workbook):
        self.workbook: Workbook = workbook
        self.xl_log: Worksheet = workbook['AIR DG SALT LOG']
        self.employee_list_start: tuple = self.find_first_employee()
        self.employee_list: list = self.get_employee_list()
        self.pcms: dict = self.get_pcm_list()
        self.drill_sheets = self.get_supp_drills()
        self.week_row: int = self.get_week_row()
        self.week_cols: list = self.get_week_cols(self.week_row)

        self.monthly_drill_date_col = None
        self.monthly_drill_result_col = None
        self._get__monthly_training_drill_cols()

        self.operation_name_cell = self._get_operation_cell()
        self.operation_name = self.operation_name_cell.value

        self.weeks = list()
        for week_col in self.week_cols:
            self.weeks.append(SaltWeek(log=self.xl_log, start_row=self.week_row, start_col=week_col))

        # Set supplemental drill sheet # and correct PCM topic info on Week objects
        for week in self.weeks:
            week.set_supp_drill(self.drill_sheets)
            week.set_correct_PCM(self.pcms)

    def set_highlight(self, cell: Cell) -> None:
        cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFF200', type='rgb'),
                                bgColor=Color(rgb='FFFFFF00', type='rgb'))

    def find_first_employee(self) -> tuple:
        for cell in self.xl_log.iter_rows(min_col=2, max_col=2):
            try:
                if cell[0].value.lower() == 'employee name':
                    return cell[0].column, cell[0].row + 1
            except AttributeError:  # Merged cells don't have value attribute
                pass                # and NoneType doesn't have to_lower()

    def get_employee_list(self):
        ee_list = list()
        for cell in self.xl_log.iter_rows(min_col=self.employee_list_start[0],
                                          max_col=self.employee_list_start[0],
                                          min_row=self.employee_list_start[1]):

            # The employee slots end at a merged cell and should never be merged cells themselves
            if isinstance(cell[0], MergedCell):
                break
            # If the line is blank, skip it
            if cell[0].value is None or cell[0].value.strip() == '':
                continue
            # Otherwise, add the employee to the list
            ee_list.append(Employee(cell[0].value.strip(), cell[0]))

        return ee_list

    def get_pcm_topic(self, sheet: Worksheet) -> str:
        for row in sheet.iter_rows(min_col=1, max_col=15, max_row=5):
            for cell in row:
                if cell.value is not None:
                    return cell.value

        raise Exception('No PCM topic found in cells searched')

    def get_pcm_list(self) -> dict:
        pcm_details = dict()
        pcm_sheets = [item for item in self.workbook.sheetnames if item.strip().startswith('PCM')]

        for item in pcm_sheets:
            pcm_date = self._parse_date(item)
            topic = self.get_pcm_topic(self.workbook[item])
            pcm_details[pcm_date] = topic

        return pcm_details

    def get_supp_drills(self) -> dict:
        supp_drills = dict()
        supp_drill_sheets = [item for item in self.workbook.sheetnames if 'drill' in item.strip().lower()]

        for item in supp_drill_sheets:
            drill_date = self._parse_date(item)
            drill_sheet = self._find_drill_sheet_name(self.workbook[item])
            supp_drills[drill_date] = drill_sheet

        return supp_drills

    def _find_drill_sheet_name(self, sheet: Worksheet) -> str:
        for row in sheet.iter_rows(max_col=15):
            for cell in row:
                try:
                    if cell.value != None: return cell.value
                except AttributeError:      #MergedCells have no value attribute
                    pass

    def get_week_row(self) -> int:
        for row in self.xl_log.iter_rows(max_col=10):
            for cell in row:
                try:
                    if 'week' in cell.value.lower():
                        return cell.row
                except AttributeError:
                    pass
        return None

    def get_week_cols(self, week_row) -> list:
        week_cols = list()
        for cell in self.xl_log.iter_cols(min_row=week_row, max_row=week_row, max_col=30):
            try:
                if 'week' in cell[0].value.lower():
                    week_cols.append(cell[0].column)
            except AttributeError:
                pass
        return week_cols

    def _get__monthly_training_drill_cols(self):
        base_col = None
        for cell in self.xl_log.iter_cols(min_row=self.week_row, max_row=self.week_row):
            try:
                if 'monthly' in cell[0].value.lower():
                    base_col = cell[0].column
            except AttributeError:
                pass
            if base_col is not None:
                break

        for row in self.xl_log.iter_rows(min_row=self.week_row, min_col=base_col, max_col=base_col+1):
            for cell in row:
                try:
                    if 'date' in cell.value.lower():
                        self.monthly_drill_date_col = cell.column
                    if 'result' in cell.value.lower():
                        self.monthly_drill_result_col = cell.column
                except AttributeError:
                    pass
            if (self.monthly_drill_date_col is not None) and (self.monthly_drill_result_col is not None):
                break

    def _get_operation_cell(self):
        row_num = None
        col_num = None

        for row in self.xl_log.iter_rows(max_col=10):
            for cell in row:
                try:
                    if row_num is None and 'operation' in cell.value.lower():
                        row_num = cell.row
                        continue
                    if row_num is not None:
                        if type(cell).__name__ == 'Cell':
                            col_num = cell.column
                except AttributeError:
                    pass
                if (row_num is not None) and (col_num is not None):
                    break
            if (row_num is not None) and (col_num is not None):
                break

        return self.xl_log.cell(row=row_num, column=col_num)

    def _parse_date(self, _string:str) -> date:
        date_string = re.search(r'\d{1,2}[/-]\d{1,2}[-/]\d{2,4}', _string).group(0)
        try:
            date = datetime.strptime(date_string, '%m-%d-%Y')
        except ValueError:
            try:
                date = datetime.strptime(date_string, '%m/%d/%Y')
            except ValueError:
                raise Exception ('Could not parse date')

        return date.date()











            

